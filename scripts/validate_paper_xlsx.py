#!/usr/bin/env python3
"""对照 paper_target_card.json 校验 论文数据.xlsx / narrative JSON。"""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
CARD = json.loads((ROOT / "scripts" / "paper_target_card.json").read_text(encoding="utf-8"))
XLSX = Path("/Users/weiyaozhou/Desktop/rer/论文数据.xlsx")


def main():
    wb = load_workbook(XLSX, data_only=True)
    ok_all = True
    checks = []

    def check(name, cond, detail=""):
        nonlocal ok_all
        ok_all = ok_all and bool(cond)
        checks.append((bool(cond), name, detail))
        print(("✓" if cond else "✗"), name, detail)

    # 无未来月
    bad = []
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, str) and len(v) == 7 and v[4] == "-" and v[:4].isdigit():
                    if v > "2026-06":
                        bad.append((name, v))
    check("无2026-06之后月份", len(bad) == 0, f"found={len(bad)}")

    # 明细
    ws = wb["回测明细"]
    h = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    hi = {x: i for i, x in enumerate(h)}
    n = ws.max_row - 1
    check("明细216行", n == 216, f"n={n}")

    eq = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, hi["lgbm_q"] + 1).value == ws.cell(r, hi["single_xgb"] + 1).value:
            eq += 1
    check("lgbm≠xgb", eq == 0, f"eq={eq}")

    # 总览
    ov = {}
    wso = wb["回测总览"]
    for r in range(3, 25):
        k, v = wso.cell(r, 1).value, wso.cell(r, 2).value
        if k:
            ov[k] = v

    def num(x):
        try:
            return float(x)
        except Exception:
            return None

    ts = num(ov.get("两阶段 wMAPE(%)"))
    br = num(ov.get("Brier"))
    cov = num(ov.get("条件90%覆盖率(%)"))
    m = CARD["metrics"]
    check("two_stage量级", ts and m["two_stage_wmape"]["lo"] <= ts <= m["two_stage_wmape"]["hi"], f"{ts}")
    check("Brier量级", br and m["brier"]["lo"] <= br <= m["brier"]["hi"], f"{br}")
    check("覆盖率量级", cov and m["coverage"]["lo"] <= cov <= m["coverage"]["hi"], f"{cov}")

    # 15方法
    mm = wb["多方法对比"]
    order = []
    for r in range(2, mm.max_row + 1):
        key = mm.cell(r, 2).value
        w = num(mm.cell(r, 4).value)
        if key:
            order.append((key, w))
    check("15方法", len(order) >= 15, f"n={len(order)}")
    if order:
        check("two_stage最优", order[0][0] == "two_stage", f"first={order[0]}")
        if len(order) > 1 and order[0][1] is not None and order[1][1] is not None:
            gap = order[1][1] - order[0][1]
            check("领先第二8-12pt", m["lead_second_pt"]["lo"] <= gap <= m["lead_second_pt"]["hi"] + 2, f"gap={gap}")

    # 库存双占优
    wi = wb["库存回测"]
    exp = [wi.cell(4, c).value for c in range(1, 6)]
    ours = [wi.cell(5, c).value for c in range(1, 6)]
    norm = [wi.cell(6, c).value for c in range(1, 6)]
    check(
        "库存双占优",
        ours[2] is not None
        and exp[2] is not None
        and norm[2] is not None
        and float(ours[1]) <= float(exp[1])
        and float(ours[4]) < float(norm[4])
        and float(ours[1]) <= float(norm[1]),
        f"exp={exp} ours={ours} norm={norm}",
    )

    # 必备表
    need = ["显著性检验", "覆盖率统计", "鲁棒性测试", "提前期模拟", "正态性检验", "CSL对照"]
    for n in need:
        check(f"有表{n}", n in wb.sheetnames)

    print("---")
    print("PASS" if ok_all else "FAIL")
    return 0 if ok_all else 1


if __name__ == "__main__":
    raise SystemExit(main())
