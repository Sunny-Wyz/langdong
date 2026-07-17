#!/usr/bin/env python3
"""将 paper_narrative_result.json + DB 消耗导出为 Desktop/rer/论文数据.xlsx"""
from __future__ import annotations

import json
import shutil
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pymysql
from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[1]
JSON_PATH = ROOT / "scripts" / "paper_narrative_result.json"
OUT = Path("/Users/weiyaozhou/Desktop/rer/论文数据.xlsx")
LABELS = ROOT / "sql" / ".paper_part_labels.json"

METHODS = [
    "two_stage", "lgbm_q", "single_xgb", "ngboost", "tft", "deepar",
    "nhits", "rf", "mapa", "adida", "tsb", "croston", "sba", "es", "sma3",
]


def r2(x):
    try:
        return round(float(x), 2)
    except Exception:
        return x


def load_demand():
    conn = pymysql.connect(host="127.0.0.1", user="admin", password="123456", database="spare_db", charset="utf8mb4")
    dem = defaultdict(dict)
    names = {}
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT code, name FROM spare_part")
            for code, name in cur.fetchall():
                names[str(code)] = name
            cur.execute(
                """
                SELECT sp.code, DATE_FORMAT(r.approve_time,'%Y-%m'), CAST(SUM(ri.out_qty) AS SIGNED)
                FROM biz_requisition_item ri
                JOIN biz_requisition r ON ri.req_id=r.id
                JOIN spare_part sp ON ri.spare_part_id=sp.id
                WHERE r.req_status IN ('OUTBOUND','INSTALLED') AND ri.out_qty>0
                  AND DATE_FORMAT(r.approve_time,'%Y-%m') <= '2026-06'
                GROUP BY sp.code, DATE_FORMAT(r.approve_time,'%Y-%m')
                """
            )
            for code, month, qty in cur.fetchall():
                dem[str(code)][str(month)] = float(qty)
    finally:
        conn.close()
    return dem, names


def months_between(a="2023-01", b="2026-06"):
    y, m = int(a[:4]), int(a[5:7])
    ye, me = int(b[:4]), int(b[5:7])
    out = []
    while (y, m) <= (ye, me):
        out.append(f"{y:04d}-{m:02d}")
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return out


def write_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
        ws.cell(1, c).font = Font(bold=True)
    for i, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            ws.cell(i, c, v)
    return ws


def main():
    data = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    dem, names = load_demand()
    labels_meta = json.loads(LABELS.read_text(encoding="utf-8")) if LABELS.exists() else {"labels": {}}
    labels = labels_meta.get("labels", {})
    months = months_between()
    parts36 = data["parts"]
    detail = data["detail"]

    if OUT.exists():
        bak = OUT.with_name(f"论文数据_备份_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        shutil.copy2(OUT, bak)
        print("backup", bak)

    wb = Workbook()
    # 说明
    ws = wb.active
    ws.title = "说明"
    o = data["overall"]
    lines = [
        "论文数据（系统重跑输出 + 叙事回测结果）",
        "数据来源：spare_db 领用出库（PAPER_REPRO_SEED v6）",
        "",
        "月份跨度：2023-01 ～ 2026-06（42 个月）",
        f"测试窗：{','.join(data['testMonths'])}",
        f"备件数（回测）：{data['partCount']}；样本数：{data['sampleCount']}",
        f"正需求点数：{o.get('cov_positivePoints')}；条件覆盖：{o.get('cov_covered')}/{o.get('cov_positivePoints')} = {o.get('cov_coverageRate')}%",
        f"两阶段 wMAPE：{o.get('wmapeTwoStage')}%；Brier：{o.get('brier')}",
        f"代表件 FOCUS：{data.get('focusPartCode')}",
        "",
        "工作表：说明 | 分层标签 | 月度消耗明细 | 备件×月份矩阵 | 备件汇总 | 回测总览 | 多方法对比 | 分层指标 | 代表件序列 | 库存回测 | 扩展实验 | 回测明细 | 显著性检验 | 覆盖率统计 | 鲁棒性测试 | 提前期模拟 | 正态性检验 | CSL对照",
        "",
        "口径说明：",
        "1) 回测结果为系统运行输出，非 PDF 静态抄录；2026-01～06 实际值取自原始消耗数据。",
        "2) wMAPE=Σ|实际−预测|/Σ实际×100；Brier=mean((p−1{y>0})²)；条件覆盖率=正需求点中 L≤实际≤U 的比例。",
        "3) MASE 分母为各备件训练期 naive 一步 MAE（明细 naiveMae 列）。",
        "4) CRPS：两阶段按零膨胀 Gamma 混合近似（p、k、μ）；其他为同趋势代理。",
        "5) 鲁棒性噪声行属独立运行，不可由本簿复算；零占比行可由明细子集复算。",
        "6) 随机种子 20260518；库存 MC 3000 次；截止月 2026-06。",
    ]
    for i, line in enumerate(lines, 1):
        ws.cell(i, 1, line)
        if i == 1:
            ws.cell(i, 1).font = Font(bold=True, size=12)

    # 分层标签
    lab_rows = []
    for code, info in sorted(labels.items()):
        lab_rows.append([code, names.get(code, ""), info.get("abc"), info.get("xyz"), info.get("combo")])
    write_sheet(wb, "分层标签", ["备件编码", "备件名称", "ABC", "XYZ", "组合"], lab_rows)

    # 月度消耗明细
    long_rows = []
    for code in sorted(dem.keys()):
        abc = labels.get(code, {}).get("abc", "")
        xyz = labels.get(code, {}).get("xyz", "")
        combo = labels.get(code, {}).get("combo", (abc or "") + (xyz or ""))
        for m in months:
            q = dem[code].get(m, 0)
            if q > 0:
                long_rows.append([None, code, names.get(code, ""), m, q, abc, xyz, combo])
    write_sheet(
        wb,
        "月度消耗明细",
        ["备件ID", "备件编码", "备件名称", "月份", "出库数量", "ABC", "XYZ", "组合"],
        long_rows,
    )

    # 矩阵
    mh = ["备件编码", "备件名称", "ABC", "XYZ", "组合"] + months + ["合计", "非零月数", "零月占比%"]
    mrows = []
    for code in sorted(dem.keys()):
        abc = labels.get(code, {}).get("abc", "")
        xyz = labels.get(code, {}).get("xyz", "")
        combo = labels.get(code, {}).get("combo", "")
        series = [dem[code].get(m, 0) for m in months]
        total = sum(series)
        nz = sum(1 for v in series if v > 0)
        z = len(series) - nz
        mrows.append([code, names.get(code, ""), abc, xyz, combo] + series + [total, nz, r2(100 * z / len(series))])
    write_sheet(wb, "备件×月份矩阵", mh, mrows)

    # 备件汇总
    srows = []
    for code in sorted(dem.keys()):
        abc = labels.get(code, {}).get("abc", "")
        xyz = labels.get(code, {}).get("xyz", "")
        combo = labels.get(code, {}).get("combo", "")
        series = [dem[code].get(m, 0) for m in months]
        pos = [v for v in series if v > 0]
        nz = len(pos)
        z = len(series) - nz
        srows.append(
            [
                code,
                names.get(code, ""),
                abc,
                xyz,
                combo,
                sum(series),
                nz,
                z,
                r2(100 * z / len(series)),
                r2(sum(pos) / len(pos)) if pos else 0,
                max(series) if series else 0,
            ]
        )
    write_sheet(
        wb,
        "备件汇总",
        ["备件编码", "备件名称", "ABC", "XYZ", "组合", "总出库量", "非零月数", "零月数", "零月占比%", "正需求均值", "最大月需求"],
        srows,
    )

    # 回测总览
    ov = data["overall"]
    paper = data.get("paperTargets", {})
    overview = [
        ["指标", "数值", "论文参照（量级）"],
        ["两阶段 wMAPE(%)", ov.get("wmapeTwoStage"), paper.get("wmape")],
        ["SMA-3 wMAPE(%)", ov.get("wmapeSma3"), 48.83],
        ["优于 SMA (pt)", data.get("advantageOverSma"), "≥8"],
        ["单阶段 XGB wMAPE(%)", ov.get("wmapeSingleXgb"), 27.59],
        ["RF wMAPE(%)", data["overallMethods"].get("rf"), 33.75],
        ["SBA wMAPE(%)", data["overallMethods"].get("sba"), "—"],
        ["Croston wMAPE(%)", data["overallMethods"].get("croston"), "—"],
        ["LGBM 分位数 wMAPE(%)", data["overallMethods"].get("lgbm_q"), 22.9],
        ["Brier", ov.get("brier"), paper.get("brier")],
        ["条件90%覆盖率(%)", ov.get("cov_coverageRate"), paper.get("coverage90")],
        ["正需求点数", ov.get("cov_positivePoints"), 154],
        ["覆盖点数", ov.get("cov_covered"), 140],
        ["样本数", data["sampleCount"], 216],
        ["备件数", data["partCount"], 36],
        ["代表件", data.get("focusPartCode"), "表3-4角色"],
        ["测试月份", ",".join(data["testMonths"]), "滚动6月"],
    ]
    ws = wb.create_sheet("回测总览")
    ws["A1"] = "真实实验叙事回测结果（运行输出）"
    ws["A1"].font = Font(bold=True, size=12)
    for i, row in enumerate(overview, 3):
        for c, v in enumerate(row, 1):
            ws.cell(i, c, v)
            if i == 3:
                ws.cell(i, c).font = Font(bold=True)

    # 多方法
    mm_rows = []
    for t in data["table36"]:
        mm_rows.append(
            [
                t["method"],
                t["methodKey"],
                t["category"],
                t["wmape"],
                t.get("mase"),
                t.get("crps"),
                t.get("coverage") if t.get("coverage") is not None else "—",
                t.get("brier") if t.get("brier") is not None else "—",
                "是" if t.get("probabilistic") else "否",
            ]
        )
    write_sheet(
        wb,
        "多方法对比",
        ["方法", "方法键", "类别", "整体wMAPE(%)", "MASE", "CRPS", "条件覆盖率(%)", "Brier", "是否概率分布"],
        mm_rows,
    )

    # 分层
    layer_rows = []
    for item in data.get("byAbc", []):
        layer_rows.append(
            [
                "ABC",
                item.get("group"),
                item.get("n"),
                item.get("wmape_two_stage"),
                item.get("wmape_sma3"),
                item.get("wmape_single_xgb"),
                item.get("wmape_rf"),
                item.get("wmape_sba"),
                item.get("brier"),
                item.get("cov_coverageRate"),
            ]
        )
    for item in data.get("byXyz", []):
        layer_rows.append(
            [
                "XYZ",
                item.get("group"),
                item.get("n"),
                item.get("wmape_two_stage"),
                item.get("wmape_sma3"),
                item.get("wmape_single_xgb"),
                item.get("wmape_rf"),
                item.get("wmape_sba"),
                item.get("brier"),
                item.get("cov_coverageRate"),
            ]
        )
    for item in data.get("byMonth", []):
        layer_rows.append(
            [
                "按月",
                item.get("month"),
                item.get("n"),
                item.get("wmape_two_stage"),
                item.get("wmape_sma3"),
                item.get("wmape_single_xgb"),
                item.get("wmape_rf"),
                item.get("wmape_sba"),
                "",
                "",
            ]
        )
    write_sheet(
        wb,
        "分层指标",
        ["维度", "分组", "样本数", "两阶段wMAPE", "SMA3", "单阶段XGB", "RF", "SBA", "Brier", "条件覆盖率"],
        layer_rows,
    )

    # 代表件
    focus = data.get("focusPartCode")
    fs = data.get("focusSeries") or []
    ws = wb.create_sheet("代表件序列")
    ws["A1"] = "代表件"
    ws["B1"] = focus
    ws["C1"] = "（论文表3-4角色）"
    ws["A2"] = "方法wMAPE"
    fw = data.get("focusWmape") or {}
    for j, m in enumerate(METHODS):
        ws.cell(2, j + 2, m)
        ws.cell(3, j + 2, fw.get(m))
    headers = ["月份", "实际"] + METHODS + ["发生概率p", "L", "U"]
    for c, h in enumerate(headers, 1):
        ws.cell(5, c, h)
        ws.cell(5, c).font = Font(bold=True)
    for i, r in enumerate(fs):
        vals = [r["month"], r["actual"]] + [r["preds"].get(m) for m in METHODS] + [
            r.get("occurrenceProb"),
            r.get("lowerBound"),
            r.get("upperBound"),
        ]
        for c, v in enumerate(vals, 1):
            ws.cell(6 + i, c, v)

    # 库存
    inv = data["inventory"]
    ws = wb.create_sheet("库存回测")
    ws["A1"] = inv.get("note", "")
    ws["A3"] = "汇总方法"
    for c, h in enumerate(["汇总方法", "缺货月次", "缺货量", "满足率%", "平均月末库存"], 1):
        ws.cell(3, c, h)
        ws.cell(3, c).font = Font(bold=True)
    for i, s in enumerate(inv.get("summary", []), 4):
        ws.cell(i, 1, s["method"])
        ws.cell(i, 2, s["stockoutMonths"])
        ws.cell(i, 3, s["stockoutQty"])
        ws.cell(i, 4, s["fillRate"])
        ws.cell(i, 5, s["avgInv"])
    be = inv.get("breakEven", {})
    ws["A8"] = "盈亏平衡"
    ws["A8"].font = Font(bold=True)
    ws["A9"] = "新增持有量(件·月)"
    ws["B9"] = be.get("extraHolding")
    ws["A10"] = "缺货削减量(件)"
    ws["B10"] = be.get("stockoutCut")
    ws["A11"] = "盈亏平衡阈值(月)"
    ws["B11"] = be.get("months")
    for c, h in enumerate(["组合", "备件", "方法", "ROP", "缺货月", "缺货量", "满足率%", "均库存"], 1):
        ws.cell(13, c, h)
        ws.cell(13, c).font = Font(bold=True)
    ri = 14
    for item in inv.get("byCombo", []):
        for m in item["methods"]:
            ws.cell(ri, 1, item["combo"])
            ws.cell(ri, 2, item["partCode"])
            ws.cell(ri, 3, m["method"])
            ws.cell(ri, 4, m["rop"])
            ws.cell(ri, 5, m["stockoutMonths"])
            ws.cell(ri, 6, m["stockoutQty"])
            ws.cell(ri, 7, m["fillRate"])
            ws.cell(ri, 8, m["avgInv"])
            ri += 1

    # 扩展实验
    ws = wb.create_sheet("扩展实验")
    ws["A1"] = "消融"
    ws["A1"].font = Font(bold=True)
    for c, h in enumerate(["配置", "wMAPE(%)", "相对降幅", "说明"], 1):
        ws.cell(2, c, h)
        ws.cell(2, c).font = Font(bold=True)
    for i, a in enumerate(data.get("ablation", []), 3):
        if a.get("wmape") is None and a.get("config") == "口径备注":
            ws.cell(i, 1, a.get("note"))
            continue
        ws.cell(i, 1, a.get("config"))
        ws.cell(i, 2, a.get("wmape"))
        ws.cell(i, 3, a.get("delta"))
        ws.cell(i, 4, a.get("note"))
    ws["A9"] = "k策略"
    ws["A9"].font = Font(bold=True)
    for c, h in enumerate(["方案", "k明细", "整体覆盖率%", "Z类覆盖率%", "Brier", "说明"], 1):
        ws.cell(10, c, h)
        ws.cell(10, c).font = Font(bold=True)
    for i, k in enumerate(data.get("kStrategy", []), 11):
        ws.cell(i, 1, k.get("scheme"))
        ws.cell(i, 2, json.dumps(k.get("k"), ensure_ascii=False) if not isinstance(k.get("k"), (int, float)) else k.get("k"))
        ws.cell(i, 3, k.get("coverage"))
        ws.cell(i, 4, k.get("zCoverage"))
        ws.cell(i, 5, k.get("brier"))
        ws.cell(i, 6, k.get("note"))
    ws["A15"] = "相对基线优势"
    ws["A15"].font = Font(bold=True)
    for c, h in enumerate(["对比方法", "备件数", "两阶段更优数", "均优势pt", "显著占优"], 1):
        ws.cell(16, c, h)
        ws.cell(16, c).font = Font(bold=True)
    for i, s in enumerate(data.get("significance", []), 17):
        ws.cell(i, 1, s.get("vs"))
        ws.cell(i, 2, s.get("parts"))
        ws.cell(i, 3, s.get("twoStageBetterCount"))
        ws.cell(i, 4, s.get("meanGapWmape"))
        ws.cell(i, 5, "是" if s.get("significant") else "否")

    # 回测明细
    dheaders = (
        ["备件编码", "月份", "ABC", "XYZ", "实际"]
        + METHODS
        + ["p", "L", "U", "k", "naive_mae", "mu"]
    )
    drows = []
    for r in detail:
        drows.append(
            [r["partCode"], r["month"], r["abc"], r["xyz"], r["actual"]]
            + [r["preds"].get(m) for m in METHODS]
            + [r.get("occurrenceProb"), r.get("lowerBound"), r.get("upperBound"), r.get("k"), r.get("naiveMae"), r.get("mu")]
        )
    write_sheet(wb, "回测明细", dheaders, drows)

    # 显著性
    sig_rows = [
        [
            s.get("vs"),
            s.get("methodKey"),
            s.get("wilcoxonP"),
            s.get("holmP"),
            s.get("effectR"),
            "是" if s.get("significant") else "否",
            s.get("twoStageBetterCount"),
        ]
        for s in data.get("significance", [])
    ]
    write_sheet(
        wb,
        "显著性检验",
        ["对比方法", "方法键", "Wilcoxon_p", "Holm校正p", "效应量r", "是否显著(α=0.05)", "两阶段更优数"],
        sig_rows,
    )

    # 覆盖率统计
    cs_rows = []
    for c in data.get("coverageStats", []):
        if c.get("group") == "Z_bootstrap":
            continue
        cs_rows.append(
            [
                c.get("group"),
                c.get("positivePoints"),
                c.get("covered"),
                c.get("missed"),
                c.get("coverageRate"),
                c.get("avgWidth"),
                c.get("wilsonLow"),
                c.get("wilsonHigh"),
            ]
        )
    ws = write_sheet(
        wb,
        "覆盖率统计",
        ["分组", "正需求点数", "覆盖点数", "未覆盖数", "条件覆盖率%", "平均区间宽度", "Wilson95下限", "Wilson95上限"],
        cs_rows,
    )
    boot = next((c for c in data.get("coverageStats", []) if c.get("group") == "Z_bootstrap"), None)
    if boot:
        r0 = len(cs_rows) + 3
        ws.cell(r0, 1, "Z类bootstrap(1000次)")
        ws.cell(r0, 1).font = Font(bold=True)
        ws.cell(r0 + 1, 1, "均值%")
        ws.cell(r0 + 1, 2, boot.get("mean"))
        ws.cell(r0 + 2, 1, "标准差%")
        ws.cell(r0 + 2, 2, boot.get("sd"))
        ws.cell(r0 + 3, 1, "经验2.5%")
        ws.cell(r0 + 3, 2, boot.get("p2_5"))
        ws.cell(r0 + 4, 1, "经验97.5%")
        ws.cell(r0 + 4, 2, boot.get("p97_5"))

    # 鲁棒性
    write_sheet(
        wb,
        "鲁棒性测试",
        ["场景", "两阶段wMAPE", "单阶段wMAPE", "口径说明"],
        [[r.get("scene"), r.get("twoStage"), r.get("singleXgb"), r.get("note")] for r in data.get("robustness", [])],
    )

    # 提前期
    write_sheet(
        wb,
        "提前期模拟",
        ["备件", "组合", "角色", "E[D_L]", "σ_L", "Q0.95", "Q0.99", "ROP", "安全库存", "备注"],
        [
            [
                r.get("partCode"),
                r.get("combo"),
                r.get("role"),
                r.get("E_DL"),
                r.get("sigma_L"),
                r.get("Q95"),
                r.get("Q99"),
                r.get("ROP"),
                r.get("safetyStock"),
                r.get("note"),
            ]
            for r in data.get("leadTime", [])
        ],
    )

    # 正态性
    write_sheet(
        wb,
        "正态性检验",
        ["组合", "备件", "Shapiro-Wilk_p", "KS_p", "AD统计量", "是否拒绝正态"],
        [
            [r.get("combo"), r.get("partCode"), r.get("shapiroP"), r.get("ksP"), r.get("adStat"), "是" if r.get("rejectNormal") else "否"]
            for r in data.get("normality", [])
        ],
    )

    # CSL
    write_sheet(
        wb,
        "CSL对照",
        ["组合", "备件", "目标CSL", "回测周期数", "无缺货周期数", "实测CSL", "实测满足率%"],
        [
            [
                r.get("combo"),
                r.get("partCode"),
                r.get("targetCsl"),
                r.get("periods"),
                r.get("noStockoutPeriods"),
                r.get("realizedCsl"),
                r.get("fillRate"),
            ]
            for r in data.get("csl", [])
        ],
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print("saved", OUT, "sheets", wb.sheetnames)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
